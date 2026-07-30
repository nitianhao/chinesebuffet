#!/usr/bin/env python3
"""Prepare Search Console restaurant-detail URLs for live enrichment."""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


CITY_FACETS = {"best", "cheap", "open-now", "top-rated", "neighborhoods"}
MODEL = "deepseek/deepseek-v4-flash"
ESTIMATED_OUTPUT_TOKENS = 250


def is_restaurant_detail_url(url: str) -> bool:
    parts = [part for part in urlparse(url).path.split("/") if part]
    return len(parts) == 3 and parts[0] == "chinese-buffets" and parts[2] not in CITY_FACETS


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--out-dir", required=True, type=Path)
    parser.add_argument("--limit", default=0, type=int)
    args = parser.parse_args()

    ws = load_workbook(args.workbook, read_only=True, data_only=True)["Table"]
    rows = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        url = str(row[0]).rstrip("/") if row and row[0] else ""
        if not url or url in seen or not is_restaurant_detail_url(url):
            continue
        seen.add(url)
        rows.append(
            {
                "sampleSet": "not_indexed",
                "model": MODEL,
                "estimatedInputTokens": 0,
                "estimatedOutputTokens": ESTIMATED_OUTPUT_TOKENS,
                "estimatedCostUsd": 0,
                "source": {"url": url},
            }
        )
        if args.limit and len(rows) >= args.limit:
            break

    args.out_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = args.out_dir / "llm-seo-summary-url-seed.jsonl"
    summary_path = args.out_dir / "llm-seo-summary-url-seed-summary.json"
    jsonl_path.write_text("\n".join(json.dumps(row, ensure_ascii=False) for row in rows) + "\n")

    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": str(args.workbook),
        "restaurantDetailUrls": len(rows),
        "files": {
            "jsonl": str(jsonl_path),
            "summary": str(summary_path),
        },
    }
    summary_path.write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
