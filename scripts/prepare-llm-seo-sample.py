#!/usr/bin/env python3
"""Prepare a tiny OpenRouter SEO-summary sample payload without API calls."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


CITY_FACETS = {"best", "cheap", "open-now", "top-rated", "neighborhoods"}
MODEL = "deepseek/deepseek-v4-flash"
INPUT_PRICE_PER_M = 0.09
OUTPUT_PRICE_PER_M = 0.18
ESTIMATED_OUTPUT_TOKENS = 250
UNSAFE_MENU_INDICATORS = (
    "insufficient",
    "non-menu",
    "no menu",
    "no actual",
    "gambling",
    "cookie",
    "javascript disabled",
    "placeholder",
    "aggregator",
    "loaded",
    "not found",
)


def restaurant_parts(url: str) -> tuple[str | None, str | None]:
    parts = [part for part in urlparse(url).path.split("/") if part]
    if len(parts) == 3 and parts[0] == "chinese-buffets" and parts[2] not in CITY_FACETS:
        return parts[1], parts[2]
    return None, None


def normalize_slug(slug: str, city_slug: str) -> str:
    state = city_slug.rsplit("-", 1)[-1] if "-" in city_slug else ""
    value = slug.lower().strip("/")
    if state:
        value = re.sub(rf"-{re.escape(state)}(-\d+)?$", "", value)
    return re.sub(r"-\d+$", "", value)


def text_len(value: object) -> int:
    return len(value.strip()) if isinstance(value, str) else 0


def load_restaurant_urls(workbook_path: Path) -> list[tuple[str, str, str]]:
    ws = load_workbook(workbook_path, read_only=True, data_only=True)["Table"]
    urls: list[tuple[str, str, str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        url = str(row[0]).rstrip("/")
        city_slug, buffet_slug = restaurant_parts(url)
        if city_slug and buffet_slug:
            urls.append((url, city_slug, buffet_slug))
    return urls


def build_buffet_index(data_path: Path) -> dict[tuple[str, str], dict]:
    by_city = json.loads(data_path.read_text())
    index: dict[tuple[str, str], dict] = {}
    for city_slug, city in by_city.items():
        for buffet in city.get("buffets", []):
            slug = buffet.get("slug") or ""
            if not slug:
                continue
            index[(city_slug, slug)] = buffet
            index[(city_slug, normalize_slug(slug, city_slug))] = buffet
    return index


def build_cuisine_index(cuisine_path: Path | None) -> dict[str, dict]:
    if not cuisine_path or not cuisine_path.exists():
        return {}
    rows = json.loads(cuisine_path.read_text())
    index: dict[str, dict] = {}
    if not isinstance(rows, list):
        return index
    for row in rows:
        place_id = row.get("placeId")
        analysis = row.get("analysis")
        if place_id and isinstance(analysis, dict):
            index[place_id] = analysis
    return index


def find_buffet(index: dict[tuple[str, str], dict], city_slug: str, buffet_slug: str) -> dict | None:
    return index.get((city_slug, buffet_slug)) or index.get((city_slug, normalize_slug(buffet_slug, city_slug)))


def summarize_hours(hours: object) -> str | None:
    if not isinstance(hours, list) or not hours:
        return None
    if len(hours) >= 7:
        return "Hours listed for all 7 days"
    return f"Hours listed for {len(hours)} days"


def source_payload(url: str, buffet: dict, cuisine_index: dict[str, dict]) -> dict:
    address = buffet.get("address") if isinstance(buffet.get("address"), dict) else {}
    city = buffet.get("cityName") or address.get("city") or ""
    state = buffet.get("stateAbbr") or address.get("state") or buffet.get("state") or ""
    existing_description = (
        buffet.get("description2")
        or buffet.get("description")
        or buffet.get("what_customers_are_saying_seo")
    )
    cuisine = cuisine_index.get(buffet.get("placeId") or "") or {}
    cuisine_indicators = cuisine.get("keyIndicators")
    if not isinstance(cuisine_indicators, list):
        cuisine_indicators = []
    menu_items = [
        str(item)
        for item in cuisine_indicators[:6]
        if item and not any(term in str(item).lower() for term in UNSAFE_MENU_INDICATORS)
    ]
    cuisine_type = cuisine.get("cuisineType")
    if menu_items and cuisine_type and cuisine_type != "Unknown":
        menu_items.insert(0, f"Cuisine analysis: {cuisine_type}")

    return {
        "url": url,
        "name": buffet.get("name"),
        "city": city,
        "state": state,
        "neighborhood": buffet.get("neighborhood"),
        "address": address.get("full") or address.get("street"),
        "rating": buffet.get("rating"),
        "reviewsCount": buffet.get("reviewsCount") or len(buffet.get("reviews") or []),
        "categories": buffet.get("categories") or [buffet.get("categoryName")] if buffet.get("categoryName") else [],
        "hoursSummary": summarize_hours(buffet.get("hours")),
        "price": buffet.get("price"),
        "hasWebsite": bool(buffet.get("website")),
        "hasPhone": bool(buffet.get("phone")),
        "photosCount": buffet.get("imagesCount") or len(buffet.get("images") or []) or len(buffet.get("imageUrls") or []),
        "existingDescription": existing_description,
        "reviewThemes": [],
        "menuItems": menu_items,
        "amenities": [],
        "nearbyContext": [],
    }


def priority_score(payload: dict) -> tuple[int, int, int]:
    reviews = int(payload.get("reviewsCount") or 0)
    desc_len = text_len(payload.get("existingDescription"))
    photos = int(payload.get("photosCount") or 0)
    weak_description = 1 if desc_len < 80 else 0
    has_core = 1 if payload.get("name") and payload.get("city") and payload.get("state") else 0
    return (weak_description, has_core + int(photos > 0), reviews)


def estimate_tokens(system_message: str, user_template: str, payload: dict) -> tuple[int, int, float]:
    prompt = f"{system_message}\n\n{user_template}\n\n{json.dumps(payload, ensure_ascii=False)}"
    input_tokens = max(1, round(len(prompt) / 4))
    total_cost = (input_tokens * INPUT_PRICE_PER_M / 1_000_000) + (
        ESTIMATED_OUTPUT_TOKENS * OUTPUT_PRICE_PER_M / 1_000_000
    )
    return input_tokens, ESTIMATED_OUTPUT_TOKENS, total_cost


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--not-indexed", required=True, type=Path)
    parser.add_argument("--indexed", required=True, type=Path)
    parser.add_argument("--not-indexed-limit", default=5, type=int)
    parser.add_argument("--indexed-limit", default=5, type=int)
    parser.add_argument("--data", default=Path("data/buffets-by-city.json"), type=Path)
    parser.add_argument("--cuisine-data", default=Path("data/cuisine-analysis-results.json"), type=Path)
    parser.add_argument("--out-dir", default=Path("scripts/output"), type=Path)
    args = parser.parse_args()

    system_message = (
        "You write grounded local restaurant summaries for a Chinese buffet directory. "
        "Use only the facts provided in the input. Return only valid JSON."
    )
    user_template = (
        "Generate one grounded restaurant SEO summary. Return JSON with llmSeoSummary, "
        "sourceFieldsUsed, omittedBecauseMissing, and riskNotes."
    )

    index = build_buffet_index(args.data)
    cuisine_index = build_cuisine_index(args.cuisine_data)
    samples = []
    missing = Counter()

    workbook_configs = [
        ("not_indexed", args.not_indexed, args.not_indexed_limit),
        ("indexed", args.indexed, args.indexed_limit),
    ]

    for label, workbook, sample_limit in workbook_configs:
        if sample_limit <= 0:
            continue
        candidates = []
        for url, city_slug, buffet_slug in load_restaurant_urls(workbook):
            buffet = find_buffet(index, city_slug, buffet_slug)
            if not buffet:
                missing[label] += 1
                continue
            payload = source_payload(url, buffet, cuisine_index)
            candidates.append((priority_score(payload), payload))
        candidates.sort(key=lambda item: item[0], reverse=True)
        for _, payload in candidates[:sample_limit]:
            input_tokens, output_tokens, estimated_cost = estimate_tokens(system_message, user_template, payload)
            samples.append(
                {
                    "sampleSet": label,
                    "model": MODEL,
                    "estimatedInputTokens": input_tokens,
                    "estimatedOutputTokens": output_tokens,
                    "estimatedCostUsd": round(estimated_cost, 6),
                    "source": payload,
                }
            )

    args.out_dir.mkdir(parents=True, exist_ok=True)
    jsonl_path = args.out_dir / "llm-seo-summary-sample.jsonl"
    summary_path = args.out_dir / "llm-seo-summary-sample-summary.json"

    jsonl_path.write_text("\n".join(json.dumps(item, ensure_ascii=False) for item in samples) + "\n")

    total_input = sum(item["estimatedInputTokens"] for item in samples)
    total_output = sum(item["estimatedOutputTokens"] for item in samples)
    total_cost = sum(item["estimatedCostUsd"] for item in samples)
    summary = {
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "model": MODEL,
        "sampleCount": len(samples),
        "missingStaticMatches": dict(missing),
        "estimatedInputTokens": total_input,
        "estimatedOutputTokens": total_output,
        "estimatedCostUsd": round(total_cost, 6),
        "files": {
            "jsonl": str(jsonl_path),
            "summary": str(summary_path),
        },
    }
    summary_path.write_text(json.dumps(summary, indent=2))
    print(json.dumps(summary, indent=2))


if __name__ == "__main__":
    main()
